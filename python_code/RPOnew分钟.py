import os
import re
import time
import numpy
from openpyxl import Workbook
from openpyxl.chart import LineChart,Reference
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

font = Font(name='微软雅黑',
                size=9,
                bold=False,
                italic=True,
                vertAlign=None,
                underline='single',
                strike=False,
                color='0B00EE')

time.perf_counter()
rpoxlsx = Workbook()
rpoxlsx.ws = rpoxlsx.create_sheet('汇总')
SummaryWriter = rpoxlsx['汇总']
SummaryWriter.cell(row=1,column=1,value='HUR组名')
SummaryWriter.cell(row=1,column=2,value='灾备等级')
SummaryWriter.cell(row=1,column=3,value='8点-20点RPO_max(min)')
SummaryWriter.cell(row=1,column=4,value='8点-20点RPO_ave(min)')
SummaryWriter.cell(row=1,column=5,value='20点-8点RPO_max(min)')
SummaryWriter.cell(row=1,column=6,value='20点-8点RPO_ave(min)')
SummaryWriter.cell(row=1,column=7,value='MJNL最大使用率(%)')
SummaryWriter.cell(row=1,column=8,value='MJNL容量(GB)')
SummaryWriter.cell(row=1,column=9,value='一天处于Pair时间(min)')
SummaryWriter.cell(row=1,column=10,value='演练时RPO_max(min)')
SummaryWriter.cell(row=1,column=11,value='SN')
SummaryWriter.cell(row=1,column=12,value='MJID')

def readfile(ctg):
    with open('OUTFILE_HSH/%s.txt'% ctg, 'r', encoding='utf-8') as f:
        data = f.readlines()
        ctgdata = []
        for d in data:
            if d.split()[2] == ctg:
                ctgdata.append(d.split())
        f.close()
    return ctgdata
datetime=[]
ctgname=[]
rpo=[]
Q_marker=[]
Q_cnt=[]
Pair_status=[]
MJNL_userate=[]
MJNL_size=[]
MJNL_useamount=[]
DR_level=[]
SN=[]
MJID=[]
p=0
q=0
for root, dirs, files in os.walk('OUTFILE_HSH/'):
    #print(files)
    for filename in files:
        ctg = re.split('[.]', filename)[0]
        #print(ctg)
        print('%s系统脚本抓取原数据'% ctg,readfile(ctg))
        #print(type(readfile(ctg)))
        print(len(readfile(ctg)))
        try:
            for i in range(0,len(readfile(ctg))):
                if int(readfile(ctg)[i+1][4]) == 0:
                    RPO = 0
                else:
                    RPO=(int(readfile(ctg)[i+1][1])-int(readfile(ctg)[i][1]))/(int(readfile(ctg)[i][4])-int(readfile(ctg)[i+1][4])+int(readfile(ctg)[i+1][3],16)-int(readfile(ctg)[i][3],16))*int(readfile(ctg)[i+1][4])
                #print(readfile(ctg)[i+1][0],readfile(ctg)[i][2],RPO)
                datetime.append(readfile(ctg)[i+1][0])
                ctgname.append(readfile(ctg)[i][2])
                rpo.append(float('%.1f' % (RPO/60)))
                Q_marker.append(readfile(ctg)[i+1][3])
                Q_cnt.append(readfile(ctg)[i+1][4])
                Pair_status.append(int(readfile(ctg)[i+1][5]))
                MJNL_userate.append(int(readfile(ctg)[i+1][6]))
                MJNL_size.append(float('%.1f' % float(readfile(ctg)[i+1][7])))
                MJNL_useamount.append(float('%.1f' % float(readfile(ctg)[i+1][8])))
                #print(len(readfile(ctg)[i]))
                if len(readfile(ctg)[i]) == 9:
                    DR_level.append('undefined')
                else:
                    DR_level.append(readfile(ctg)[i+1][9])
                SN.append(readfile(ctg)[i+1][12])
                MJID.append(readfile(ctg)[i+1][13])
        except:
            pass
        #print(datetime)
        #print(len(datetime))
        #print(ctgname)
        #print(len(ctgname))
        try:
            print('%s系统RPO值'% ctg,rpo)
            print(max(rpo[0:int(len(rpo)/2)]))
            print(numpy.mean(rpo[0:int(len(rpo)/2)]))
            print(int(len(rpo)/2))
            print(max(rpo[int(len(rpo)/2):int(len(rpo))]))
            print(type(max(rpo[int(len(rpo)/2):int(len(rpo))])))
            print(numpy.mean(rpo[int(len(rpo)/2):int(len(rpo))]))
            print(type(numpy.mean(rpo[int(len(rpo)/2):int(len(rpo))])))
            print('%s系统全天jnl使用率'% ctg,MJNL_userate)
            print('%s系统最大jnl使用率'% ctg,max(MJNL_userate[0:int(len(MJNL_userate))]))
            print(type(max(MJNL_userate[0:int(len(MJNL_userate))])))
            print('%s系统Pair状态'% ctg,sum(Pair_status[0:int(len(Pair_status))]))
            print('%s系统Pair状态指标数量'% ctg,len(Pair_status))
        except:
            pass
        #print(len(rpo))
        urrpo=list(zip(datetime,ctgname,rpo,Q_marker,Q_cnt,Pair_status,MJNL_userate,MJNL_size,MJNL_useamount,DR_level,SN,MJID))
        #urrpo=list(zip(datetime,ctgname,rpo,Pair_status,MJNL_userate,MJNL_size,MJNL_useamount))
        print('%s系统写入报告的值'% ctg,urrpo)
        print(len(urrpo))
        r=0
        c=0

        rpoxlsx.ws = rpoxlsx.create_sheet('%s'% ctg)
        link1 = '#%s!A1' % ctg
        link2 = '#汇总!A1'
        for j in range(0,len(urrpo)+1):
            try:
                rpoxlsx.ws.cell(row=1,column=c+1,value='Date&Time')
                rpoxlsx.ws.cell(row=1,column=c+2,value='RPO(min)')
                rpoxlsx.ws.cell(row=1,column=c+3,value='Q-marker')
                rpoxlsx.ws.cell(row=1,column=c+4,value='Q-cnt')
                rpoxlsx.ws.cell(row=1,column=c+5,value='Pair状态')
                rpoxlsx.ws.cell(row=1,column=c+6,value='MJNL使用率(%)')
                rpoxlsx.ws.cell(row=1,column=c+7,value='MJNL容量(GB)')
                rpoxlsx.ws.cell(row=1,column=c+8,value='MJNL使用量(GB)')
                rpoxlsx.ws.cell(row=1,column=c+9,value='演练时RPO(min)')
                rpoxlsx.ws.cell(row=1,column=c+10,value='灾备等级')
                rpoxlsx.ws.cell(row=r+1,column=11,value='返回汇总')
                rpoxlsx.ws.cell(row=r+1,column=11).font=font

                rpoxlsx.ws.cell(row=r+2,column=1,value=urrpo[j][0])
                rpoxlsx.ws.cell(row=r+2,column=c+2,value=urrpo[j][2])
                rpoxlsx.ws.cell(row=r+2,column=c+3,value=urrpo[j][3])
                rpoxlsx.ws.cell(row=r+2,column=c+4,value=int(urrpo[j][4]))
                rpoxlsx.ws.cell(row=r+2,column=c+5,value=urrpo[j][5])
                rpoxlsx.ws.cell(row=r+2,column=c+6,value=int(urrpo[j][6]))
                rpoxlsx.ws.cell(row=r+2,column=c+7,value=float(urrpo[j][7]))
                rpoxlsx.ws.cell(row=r+2,column=c+8,value=float('%.1f' % (float(urrpo[j][8]))))
                rpoxlsx.ws.cell(row=r+2,column=c+9,value=float('%.1f' % (float(urrpo[j][8])/0.4/60)))
                rpoxlsx.ws.cell(row=r+2,column=c+10,value=urrpo[j][9])
                rpoxlsx.ws.cell(row=r+1, column=11).hyperlink = (link2)
                #SummaryWriter.cell(row=p+2,column=2,value=urrpo[j][9])
                #SummaryWriter.cell(row=p+2,column=8,value=urrpo[j][7])
                #print(urrpo[j][0],urrpo[j][1],urrpo[j][2],urrpo[j][6],type(urrpo[j][6]))
                r += 1
                if urrpo[j][1] != urrpo[j+1][1]:
                    c += 1
                    r = 0
            except:
                pass
        chart1 = LineChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "HUR RPO"
        chart1.y_axis.title = 'Second'
        chart1.x_axis.title = 'Date&Time'
        data = Reference(rpoxlsx.ws, min_col=2, min_row=1, max_row=len(urrpo)+1, max_col=2)
        cats = Reference(rpoxlsx.ws, min_col=1, min_row=2, max_row=len(urrpo)+1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        rpoxlsx.ws.add_chart(chart1, "M2")

        chart2 = LineChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "MJNL使用率"
        chart2.y_axis.title = 'U%'
        chart2.x_axis.title = 'Date&Time'
        data = Reference(rpoxlsx.ws, min_col=6, min_row=1, max_row=len(urrpo)+1, max_col=6)
        cats = Reference(rpoxlsx.ws, min_col=1, min_row=2, max_row=len(urrpo)+1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        chart2.shape = 4
        rpoxlsx.ws.add_chart(chart2, "M34")

        chart3 = LineChart()
        chart3.type = "col"
        chart3.style = 10
        chart3.title = "演练时预计断开时间"
        chart3.y_axis.title = 'Second'
        chart3.x_axis.title = 'Date&Time'
        data = Reference(rpoxlsx.ws, min_col=9, min_row=1, max_row=len(urrpo)+1, max_col=9)
        cats = Reference(rpoxlsx.ws, min_col=1, min_row=2, max_row=len(urrpo)+1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        chart3.shape = 4
        rpoxlsx.ws.add_chart(chart3, "M50")

        chart4 = LineChart()
        chart4.type = "col"
        chart4.style = 10
        chart4.title = "处于Pair状态"
        chart4.y_axis.title = 'Pair'
        chart4.x_axis.title = 'Date&Time'
        data = Reference(rpoxlsx.ws, min_col=5, min_row=1, max_row=len(urrpo)+1, max_col=5)
        cats = Reference(rpoxlsx.ws, min_col=1, min_row=2, max_row=len(urrpo)+1)
        chart4.add_data(data, titles_from_data=True)
        chart4.set_categories(cats)
        chart4.shape = 4
        rpoxlsx.ws.add_chart(chart4, "M18")

        '''
        SummaryWriter.cell(row=1,column=1,value='HUR组名')
        SummaryWriter.cell(row=1,column=2,value='灾备等级')
        SummaryWriter.cell(row=1,column=3,value='8点-20点RPO_max(min)')
        SummaryWriter.cell(row=1,column=4,value='8点-20点RPO_ave(min)')
        SummaryWriter.cell(row=1,column=5,value='20点-8点RPO_max(min)')
        SummaryWriter.cell(row=1,column=6,value='20点-8点RPO_ave(min)')
        SummaryWriter.cell(row=1,column=7,value='MJNL最大使用率(%)')
        SummaryWriter.cell(row=1,column=8,value='MJNL容量(GB)')
        SummaryWriter.cell(row=1,column=9,value='一天处于Pair时间(min)')
        SummaryWriter.cell(row=1,column=10,value='演练时RPO_max(min)')
        SummaryWriter.cell(row=1,column=11,value='SN')
        SummaryWriter.cell(row=1,column=12,value='MJID')
        '''


        try:
            SummaryWriter.cell(row=p+2,column=1,value=ctg)
            SummaryWriter.cell(row=p+2,column=2,value=urrpo[1][9]) #
            SummaryWriter.cell(row=p+2,column=3,value=max(rpo[0:int(len(rpo)/2)]))
            SummaryWriter.cell(row=p+2,column=4,value=float('%.1f'%(numpy.mean(rpo[0:int(len(rpo)/2)]))))
            SummaryWriter.cell(row=p+2,column=5,value=max(rpo[int(len(rpo)/2):int(len(rpo))]))
            SummaryWriter.cell(row=p+2,column=6,value=float('%.1f'%(numpy.mean(rpo[int(len(rpo)/2):int(len(rpo))]))))
            SummaryWriter.cell(row=p+2,column=7,value=max(MJNL_userate[0:int(len(MJNL_userate))]))
            SummaryWriter.cell(row=p+2,column=8,value=urrpo[1][7]) #
            SummaryWriter.cell(row=p+2,column=9,value=float('%.1f'%(sum(Pair_status[0:int(len(Pair_status))])/len(Pair_status)*1440)))
            SummaryWriter.cell(row=p+2,column=10,value=float('%.1f'%(max(MJNL_userate[0:int(len(MJNL_userate))])/100*urrpo[1][7]/0.4/60)))
            SummaryWriter.cell(row=p+2,column=11,value=urrpo[1][10]) #
            SummaryWriter.cell(row=p+2,column=12,value=urrpo[1][11]) #
            #SummaryWriter.cell(row=p+2,column=1,value=ctg)
            SummaryWriter.cell(row=p+2,column=13,value='ForDetails')
            SummaryWriter.cell(row=p+2,column=13).font = font
            SummaryWriter.cell(row=p+2,column=13).hyperlink = (link1)
        #print('p=',p)
        #print('q=',q)
            p+=1
            q+=1
        except:
            pass
        #rpoxlsx.save('OUTFILE_HSH/rpo/rpo.xlsx')
        datetime=[]
        ctgname=[]
        rpo=[]
        Q_marker=[]
        Q_cnt=[]
        Pair_status=[]
        MJNL_userate=[]
        MJNL_size=[]
        MJNL_useamount=[]
        DR_level=[]
        SN=[]
        MJID=[]
        urrpo=[]

SummaryWriter.cell(row=1,column=15,value='运行耗时 %-5.5ss'%time.perf_counter())
rpoxlsx.save('OUTFILE_HSH-rpo/rpo.xlsx')

